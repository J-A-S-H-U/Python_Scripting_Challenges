from openpyxl import load_workbook
from tabulate import tabulate

wb = load_workbook("Interface_Data.xlsx")
ws = wb.active

data = []
header = []
interface1 = []
interface2 = []
interface3 = []
interface4 = []

for row in ws.iter_rows():
    data.append(row)

header = data[0]    
interface1 = data[1] 
interface2 = data[2] 
interface3 = data[3] 
interface4 = data[4] 

t1 = tuple(interface1)
t2 = tuple(interface2)
t3 = tuple(interface3)
t4 = tuple(interface4)

interfaces = [t1,
              t2,
              t3,
              t4]
with open("table.txt","w") as file:
    file.write(tabulate(interfaces,headers=header,tablefmt="grid"))
    file.close()
